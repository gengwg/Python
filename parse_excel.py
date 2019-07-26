"""
parse a excel file, and generate a list of FQDNs from it.
"""
import openpyxl

wb = openpyxl.load_workbook('dl/myfile.xlsx')
print wb.get_sheet_names()

# meta data
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
print sheet.title
print wb.get_active_sheet()
print sheet.max_row
print sheet.max_column

## iterate part of the file
# for row in sheet.iter_rows(min_row=2, min_col=1, max_row=6, max_col=9):
#     for cell in row:
#         print cell.value,
#     print
#

## print out all cells in all rows
# for row in sheet.rows:
#     for cell in row:
#         print cell.value,
#     print

# print out fqdn for use in prometheus target file
for row in sheet.rows:
    #print '{0}.{1}'.format(row[0].value, row[-1].value)
    #print '{0}.{1}'.format(row[0].value, row[8].value)  # 7356 only
    # colum 0 is host name,  column -1 is domain name
    print  '    "{0}.{1}",'.format(row[0].value, row[-1].value)
    #print row
